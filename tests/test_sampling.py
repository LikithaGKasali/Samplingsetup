import time


import pytest
from selenium.webdriver.support.wait import WebDriverWait
import openpyxl as openpyxl
from openpyxl.packaging import workbook
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.select import Select
from selenium.webdriver.support.wait import WebDriverWait
from webdrivermanager import ChromeDriverManager
from selenium.common.exceptions import UnexpectedAlertPresentException

from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import UnexpectedAlertPresentException

from PageObject.Setup import SetupFilter

from utils.BaseClass import BaseClass


class TestSetup(BaseClass):


    def test_sampling(self, setup_b):
        logg = self.get_logger()
        logg.info("starting the sampling purpose")
        setupfilter = SetupFilter(self.driver)
        setupfilter.loginPage().click()
        Data_Dict = {}
        book = openpyxl.load_workbook("C:\\python\\BSC.xlsx")
        sheet = book.active

        for j in range(1, sheet.max_column + 1):
            Data_Dict[sheet.cell(1, j).value] = sheet.cell(2, j).value

        setupfilter.Username().send_keys(Data_Dict["username"])
        setupfilter.Password().send_keys(Data_Dict["password"])
        setupfilter.go().click()

        setupfilter.Sampling_infocenter().click()

        self.driver.switch_to.frame("MultiWindowIframe1")

        self.wait_element(self.driver, By.XPATH, "//div[.='Sampling Setup']").click()
        self.driver.switch_to.frame(self.driver.find_element(By.XPATH, "//frame[@title = 'tempFrame']"))

        setupfilter.quarter().click()
        self.SelectByVisibleText(setupfilter.quarter_drop(), Data_Dict["Quarter"].upper())

        setupfilter.year().click()
        self.SelectByVisibleText(setupfilter.year_dropdown(), str(Data_Dict['year']))

        self.driver.find_element(By.ID, "image15").click()
        self.driver.switch_to.frame("mlovPopup")
        self.wait_element(self.driver, By.NAME, "search_name").send_keys(Data_Dict["Exit Criteria"])

        self.driver.find_element(By.CLASS_NAME, "x-btn").click()
        self.wait_element(self.driver, By.CLASS_NAME, 'x-grid3-row').click()

        self.driver.switch_to.parent_frame()
        setupfilter.round1().send_keys(Data_Dict["Round 1"])
        setupfilter.round2().send_keys(Data_Dict["Round2"])
        setupfilter.round3().send_keys(Data_Dict["Round3"])

        setupfilter.hrr_round1().send_keys(Data_Dict["Round1(HRR)"])
        setupfilter.hrr_round2().send_keys(Data_Dict["Round2(HRR)"])
        setupfilter.hrr_round3().send_keys(Data_Dict["Round3(HRR)"])

        setupfilter.Enable_Grade_E_fun().click()

        gradeE = Select(setupfilter.Enable_Grade_E_dropdown_fun())
        gradeE.select_by_visible_text(Data_Dict["Enable Grade E Functionality?"])

        setupfilter.Grade_E_Round1_fun().send_keys(Data_Dict["Round 1 % (Grade E Rep)"])
        setupfilter.Grade_E_Round2_fun().send_keys(Data_Dict["Round 2 % (Grade E Rep)"])
        setupfilter.Grade_E_Round3_fun().send_keys(Data_Dict["Round 3 % (Grade E Rep)"])

        setupfilter.Sampling_start_date_fun().send_keys(Data_Dict["Sampling Start Day"])

        setupfilter.Persistency_Ratio_fun().send_keys(Data_Dict["Persistency Ratio Threshold (For HRR"])

        self.driver.find_element(By.ID, "image29").click()
        self.driver.switch_to.frame("mlovPopup")

        setupfilter.DistRep_Status_fun().send_keys(
            Data_Dict["Dist.Rep Status (As on Last Day of the Sampling Period)"])

        self.driver.find_element(By.ID, "ext-comp-1008").click()
        self.wait_element(self.driver, By.XPATH, "//input[@id='cbox0']").click()

        self.driver.find_element(By.ID, "ext-gen31").click()
        self.driver.switch_to.parent_frame()
        self.driver.find_element(By.ID, "image30").click()
        self.driver.switch_to.frame("mlovPopup")

        setupfilter.Transaction_Type_fun().send_keys(Data_Dict["Transaction Type"])

        self.driver.find_element(By.ID, "ext-comp-1008").click()
        self.wait_element(self.driver, By.ID, "cbox0").click()
        self.driver.find_element(By.ID, "ext-comp-1010").click()

        self.driver.switch_to.parent_frame()

        self.SelectByVisibleText(setupfilter.Benefit_Status_fun(), Data_Dict["Benefit Status"])

        self.driver.find_element(By.ID, "id33").send_keys(Data_Dict["Priority"])
        self.driver.find_element(By.ID, "image37").click()

        self.driver.switch_to.frame("mlovPopup")
        checkbox = self.driver.find_elements(By.XPATH, "//input[@type='checkbox']")
        for i in checkbox:
            i.click()

        self.driver.find_element(By.ID, "ext-comp-1010").click()
        self.driver.switch_to.parent_frame()
        self.driver.find_element(By.ID, "image44").click()
        self.driver.switch_to.frame("lovPopup")

        self.driver.find_element(By.ID, "search_name").send_keys(Data_Dict["Approver"])
        self.driver.find_element(By.ID, "ext-gen95").click()
        self.wait_element(self.driver, By.ID, "cbox0").click()

        self.driver.find_element(By.ID, "ext-comp-1010").click()
        self.driver.switch_to.parent_frame()
        self.driver.switch_to.parent_frame()
        self.driver.switch_to.parent_frame()
        self.driver.find_element(By.ID, "msai_toolsubmit").click()
        Alert= self. driver.switch_to.alert
        message = Alert.text
        logg.info(message)
        Alert.accept()
        logg.info("sampling setup is completed")
        # assert ("hello" in message)


